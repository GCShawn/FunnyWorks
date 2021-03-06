#!/usr/bin/env python
# coding: utf-8

# In[9]:


# -*- coding: utf-8 -*-
"""
Created on Mon Aug  3 13:26:46 2020

@author: ShawnXiong
"""
import psycopg2
import pandas as pd
import json
import shutil
import datetime
import openpyxl
from openpyxl.styles import Border,Side,Alignment,PatternFill,Font
# # 得到当前根目录
# o_path = os.getcwd() # 返回当前工作目录
# sys.path.append(r'D:\Git_re\DataAnalysis') # 添加自己指定的搜索路径
# import tool


# In[10]:


run = 0


# In[13]:


def duplicate(src=r"D:\Git_tree\DataAnalysis\tool.py", path=r"D:\Git_tree\DataAnalysis\项目里程碑进度报表\tool.py"):
    shutil.copyfile(src, path, follow_symlinks=False)
    print("成功复制到:"+path)


def setTime(time_record, date_str=''):
    if time_record == "固定时间":
        date_now = datetime.date(*map(int, date_str.split('-')))
        return date_now
    elif time_record == "当前时间":
        date_now = datetime.datetime.now().date()
        return date_now


# 修改pandas的显示效果的子函数
def floatFormat(x):
    if abs(x) >= 1e10 or 0 < abs(x) < 1e-3:
        return "%e" % x
    else:
        return "%.4f" % x


# 修改pandas的显示效果
def pandasPrettyPrinting():
    pd.set_option('display.max_rows', None)     # 解决行显示不全
    pd.set_option('display.max_columns', None)  # 解决列显示不全
    # pd.set_option('max_colwidth', 1000)         # 解决列宽不够
    # pd.set_option('display.width', 100)         # 解决列过早换行
    pd.set_option('display.float_format', floatFormat)  # 解决浮点数总是科学计数法


# 从json中获得key的所有需要的字段
def getKeyFromJson(key_path, key_test, key_orgin, key_flag):
    if (key_flag == "测试环境") or (key_flag == "真实环境"):
        if key_flag == "测试环境":
            t = open(key_path + key_test)
            f = json.load(t)
        elif key_flag == "真实环境":
            t = open(key_path + key_orgin)
            f = json.load(t)
        key = {"database": "", "user": "",
               "password": "", "host": "", "port": ""}
        key["database"] = f[0]["database"]
        key["user"] = f[0]["user"]
        key["password"] = f[0]["password"]
        key["host"] = f[0]["host"]
        key["port"] = f[0]["port"]
    return key


# 通过获得的key连接数据库
def connectDatabase(key):
    in_database = key["database"]
    in_user = key["user"]
    in_password = key["password"]
    in_host = key["host"]
    in_port = key["port"]
    conn = psycopg2.connect(database=in_database, user=in_user,
                            password=in_password, host=in_host, port=in_port)
    print("Connecting database successfully")
    cur = conn.cursor()
    return cur


# 通过（数据库连接指针，sql语句）获得数据，是getDataFromSQLList的子函数，也可以单独使用
def getDataFromSQL(cur, in_sqlText):
    cur.execute(in_sqlText)
    data = cur.fetchall()
    columnDes = cur.description  # 获取列名
    columnNames = [columnDes[i][0] for i in range(len(columnDes))]
    result_df = pd.DataFrame([list(i) for i in data], columns=columnNames)
    return result_df


# 通过（sql动态变量空间，sql指针，sql的存放位置，sql的地址组）来获得sql的内容，具体数据放置在sql动态变量空间中
def getDataFromSQLList(sql_var_names, sql_cur, sql_path, sql_array, replacing={}):
    for sql_name in sql_array:
        sql = open(sql_path + sql_name, mode='r', encoding='utf_8_sig')
        sql_text = sql.readlines()
        sql_text = "".join(sql_text)
        sql_text = sql_text.replace('\ufeff', ' ')

        if len(replacing) != 0:
            for aim in replacing.keys():
                if aim in sql_text:
                    sql_text = sql_text.replace(aim, replacing[aim])
                    
        if sql_name.split('.')[0].find('t_') != -1:
            sql_name_attach = sql_name.split('.')[0][sql_name.split('.')[0].find('t_'):]
        else:
            sql_name_attach = sql_name.split('.')[0][sql_name.split('.')[0].find('sql_of_')+7:]

        sql_var_names['df_of_' + sql_name_attach] = getDataFromSQL(sql_cur, sql_text)
        print("sql'【 %s 】'is running " % sql_name, end="")
        print("The variable name is 【 df_of_%s 】" % sql_name_attach)

        # dataframeToCSV(sql_var_names['df'+sql_name.split('.')[0][4:]], 'df'+sql_name.split('.')[0][4:]+".csv")   #保存到本地，文件的当前目录


# 通过（文件地址，表名，标题列，空值填充规则）来从文件中获得数据，适配了CSV和EXCEL
def getDataFromFile(file_path, sheet_name='Sheet1', index_col=None, na_values=['NA']):
    data_frame = pd.DataFrame()
    if file_path.split(".")[-1] == "csv":  # 对于CSV文件
        data_frame = pd.read_csv(file_path)
        print("read file %s successfully " % file_path.split("/")[-1])
    elif file_path.split(".")[-1] == "xlsx" or file_path.split(".")[-1] == "xls":  # 对于excel文件
        data_frame = pd.read_excel(file_path, sheet_name, index_col, na_values)
    return data_frame


# 通过（数据，数据地址，写入方式，列名，编码方式）来写入文件为CSV
def dataframeToCSV(data, file_path, mode="w", index=False, encoding='utf_8_sig'):
    # index:False(不显示列名),True(显示列名);encoding:'utf_8_sig'(确保中文不乱码)
    data.to_csv(file_path, mode=mode, index=index, encoding=encoding)

    # mode =
    # t	  文本模式 (默认)。
    # x	  写模式，新建一个文件，如果该文件已存在则会报错。
    # b	  二进制模式。
    # +	  打开一个文件进行更新(可读可写)。
    # U	  通用换行模式（不推荐）。
    # r	  以只读方式打开文件。文件的指针将会放在文件的开头。这是默认模式。
    # rb  以二进制格式打开一个文件用于只读。文件指针将会放在文件的开头。这是默认模式。一般用于非文本文件如图片等。
    # r+  打开一个文件用于读写。文件指针将会放在文件的开头。
    # rb+ 以二进制格式打开一个文件用于读写。文件指针将会放在文件的开头。一般用于非文本文件如图片等。
    # w	  打开一个文件只用于写入。如果该文件已存在则打开文件，并从开头开始编辑，即原有内容会被删除。如果该文件不存在，创建新文件。
    # wb  以二进制格式打开一个文件只用于写入。如果该文件已存在则打开文件，并从开头开始编辑，即原有内容会被删除。如果该文件不存在，创建新文件。一般用于非文本文件如图片等。
    # w+  打开一个文件用于读写。如果该文件已存在则打开文件，并从开头开始编辑，即原有内容会被删除。如果该文件不存在，创建新文件。
    # wb+ 以二进制格式打开一个文件用于读写。如果该文件已存在则打开文件，并从开头开始编辑，即原有内容会被删除。如果该文件不存在，创建新文件。一般用于非文本文件如图片等。
    # a	  打开一个文件用于追加。如果该文件已存在，文件指针将会放在文件的结尾。也就是说，新的内容将会被写入到已有内容之后。如果该文件不存在，创建新文件进行写入。
    # ab  以二进制格式打开一个文件用于追加。如果该文件已存在，文件指针将会放在文件的结尾。也就是说，新的内容将会被写入到已有内容之后。如果该文件不存在，创建新文件进行写入。
    # a+  打开一个文件用于读写。如果该文件已存在，文件指针将会放在文件的结尾。文件打开时会是追加模式。如果该文件不存在，创建新文件用于读写。
    # ab+ 以二进制格式打开一个文件用于追加。如果该文件已存在，文件指针将会放在文件的结尾。如果该文件不存在，创建新文件用于读写。


# 定制化操作
def dataframeCompress(df, CombineOfUnit, UnitToDo):
    # df：           传入的dataframe
    # CombineOfUnit：这一个list内容的组合为最小单位
    # UnitToDo：     需要处理的列名

    # 先根据combineofunit进行order排序
    df = df.sort_values(by=CombineOfUnit, axis=0, ascending=True)  # A-Z排序

    OutPutList = []
    i = 0
    tmp = dict(df.iloc[i])

    for i in range(1, len(df)):
        # 开始
        if i < (len(df)-1):
            flag = 1  # 默认相同，可以通过
            for j in CombineOfUnit:  # 监测状况，如果和前一条不同，就flag为0
                if df.iloc[i][j] != df.iloc[i-1][j]:
                    flag = 0  # 有不同的
                    break

            if flag == 1:
                # 前后相同的情况下，对数据进行处理，默认为累加
                tmp[UnitToDo[0]] = tmp[UnitToDo[0]] + df.iloc[i][UnitToDo[0]]
            else:
                # 前后不相同的情况下，就写入该tmp并且更新此tmp
                OutPutList.append(list(tmp.values()))
                tmp = dict(df.iloc[i])

        else:
            flag = 1
            for j in CombineOfUnit:
                if df.iloc[i][j] != df.iloc[i-1][j]:
                    flag = 0
                    break

            if flag == 1:
                tmp[UnitToDo[0]] = tmp[UnitToDo[0]] + df.iloc[i][UnitToDo[0]]
                OutPutList.append(list(tmp.values()))  # 添加了写入
            else:
                OutPutList.append(list(tmp.values()))
                tmp = dict(df.iloc[i])
                OutPutList.append(list(tmp.values()))  # 添加了写入

    df_of_OutPutList = pd.DataFrame(data=OutPutList, columns=df.columns.values)

    return df_of_OutPutList

#将Datetime改为Str
def changeDatetimeToDateStr(df,col):
    df[col] = df[col].apply(lambda x : x.strftime('%Y-%m-%d'))
    return df

#将形如2020-10-10的日期改为Date
def changeStrToDate(str):
    date = datetime.date(*map(int, date_str.split('-')))
    return date 

#修改列数据
def changeColData(df_base,col,dic):
    df_base[col] = df_base[col].apply(lambda x: dic[x])

#修改Date到Datetime
def changeDateToDatetime(date):
    datetime = pd.to_datetime(date)
    return datetime

#修改列名
def changeCol(df_base,dic,inplace = True):
    df = df_base.rename(columns = dic, inplace = inplace)
    return df

#修改指定列的列宽（使用字典
def changeColWidthByDic(ws,dic_col_width):
    for i in dic_col_width.keys():
        ws.column_dimensions[i].width = dic_col_width[i]

#从头到未修改列宽（使用列表
def changeColWidthByList(ws,lis):
    for i in range(1,len(lis)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = lis[i-1]

#给sheet绘制全框线
def drawAllaround(ws, color = '000000'):
    all_border = Border(left=Side(border_style='thin',color='000000'),
                    right=Side(border_style='thin',color='000000'),
                    top=Side(border_style='thin',color='000000'),
                    bottom=Side(border_style='thin',color='000000'))
    for row in ws.rows:
        for cell in row:
            cell.border = all_border

#绘制中间框线,输入线左一列数字
def drawLine(ws,ws_col, mid_color = 'FF0000'):
    thick_right = Border(left=Side(border_style='thin',color='000000'),
                        right=Side(border_style='thick',color=mid_color),
                        top=Side(border_style='thin',color='000000'),
                        bottom=Side(border_style='thin',color='000000'))
    
    thick_left = Border(left=Side(border_style='thick',color='000000'),
                        right=Side(border_style='thin',color=mid_color),
                        top=Side(border_style='thin',color='000000'),
                        bottom=Side(border_style='thin',color='000000'))
    for i in range(1,ws.max_row+1):
        ws[i][ws_col-1].border = thick_right
        ws[i][ws_col].border = thick_left    
            
#居中
def changeToMiddle(ws):
    for row in ws.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

#在指定位置插入列，默认为col后
def insertCol(df,col, col_insert_name, value, posi = 1):
    df.insert(list(df.columns).index(col)+posi, col_insert_name, value)
    
#字典
dic_col_1 = {'contract_gid': '合同GID', 'contract_name': '合同名称', 'contract_paydue_type': '合同收款节点类型', 'milestone_id': '合同收款节点匹配项目里程碑', 'contract_paydue_date1': '开始时间', 'contract_paydue_date2': '结束时间','add_value_addition': '合同总金额（合同金额+变更金额）', 'sum_contract_paydue_amount': '收款节点的累计金额', 'whether_equal': '是否相等', 'sum_positive_invoice_amount': '正累计开票金额', 'sum_negative_invoice_amount': '负累计开票金额', 'sum_receipts_amount': '累计收款金额'}

dic_contract_paydue = {'IMPL': '实施款', 'LICS': '软件款','MD': '人天费', 'OM': '运维费', 'SBCT': '订阅费'}

dic_col_2 = {"project_name":"项目名称","project_gid":"项目GID","project_manager_name":"项目经理姓名","project_director_name":"项目总监姓名","project_status":"项目状态","milestone_name":"里程碑名称","milestone_status":"里程碑状态","end_date":"里程碑结束日期","estimated_end_date":"里程碑预计结束日期","contract_paydue_name_agg":"对应收款节点","milestone_amount":"收款金额"}

dic_milestone_status = {"ACI":"已完成","IPR":"进行中","INIT":"未启动"}

dic_project_status = {"BPL":"计划中","CMP":"已结项","HLD":"挂起","MAINT":"运维中","QUOT":"报价中","WIP":"进行中"}

dic_milestone_name = {"PRSRT":"项目启动","ASIS":"现状调研","TOBE":"蓝图设计","DEVP":"开发构建","UAT":"UAT","PREPROD":"试运行","PROD":"上线支持","MAINT":"系统运维","FUTURE":"未来"}

            
if run:
    duplicate()

